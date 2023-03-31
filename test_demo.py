from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import global_constants

class Test_Demo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(global_constants.URL)
        self.forlder_path_of_name = str(date.today()) 
        Path(self.forlder_path_of_name).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()
    
    def visibility_element_located(self, located):
        WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((located)))

    def get_data():
        excel_file = openpyxl.load_workbook("./data/invalid_login.xlsx")
        selected_sheet = excel_file["Sheet1"]

        total_row = selected_sheet.max_row
        data = []

        for item in range(2, total_row+1):
            user_name = selected_sheet.cell(item, 1).value
            password = selected_sheet.cell(item, 2).value
            tuple_data = (user_name, password)
            data.append(tuple_data)

        return data

    @pytest.mark.parametrize("username, pasword", get_data())
    def test_invalid_login(self, username, pasword):
        self.visibility_element_located((By.ID, "user-name"))
        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys(username)

        self.visibility_element_located((By.ID, "password"))
        password = self.driver.find_element(By.ID, "password")
        password.send_keys(pasword)

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        
        self.driver.save_screenshot(self.forlder_path_of_name+f"/invalid-test-for-{username}-{pasword}.png")
        assert error_message.text == "Epic sadface: Username and password do not match any user in this service"

    def test_empty_skip(self):
        self.visibility_element_located((By.ID, "login-button"))

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()      

        self.visibility_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))

        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert error_message.text == "Epic sadface: Username is required"

        self.visibility_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3/button"))

        skip_button = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3/button")
        
        skip_button.click()

    def test_emtyp_password(self):
        self.visibility_element_located((By.ID, "user-name"))

        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys("test user name")
        password = self.driver.find_element(By.ID, "password")

        self.visibility_element_located((By.ID, "login-button"))

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        self.visibility_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))

        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert error_message.text == "Epic sadface: Password is required"
  
    def test_special_user(self):
        self.visibility_element_located((By.ID, "user-name"))
        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys("locked_out_user")

        password = self.driver.find_element(By.ID, "password")
        password.send_keys("secret_sauce")

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        self.visibility_element_located((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))

        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert error_message.text == "Epic sadface: Sorry, this user has been locked out."

    def test_route(self):
        self.visibility_element_located((By.ID, "user-name"))
        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys("standard_user")   

        password = self.driver.find_element(By.ID, "password")
        password.send_keys("secret_sauce")

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        self.visibility_element_located((By.CLASS_NAME, "inventory_item"))

        product_list = self.driver.find_elements(By.CLASS_NAME, "inventory_item")
        print(f"Products lenght : {len(product_list)}")
        assert len(product_list) 

    # Product sayfasındaki fiyatı en yüksek ürünün bulunması testi
    def test_high_product_of_page(self):
        self.visibility_element_located((By.ID, "user-name"))
        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys("standard_user")

        password = self.driver.find_element(By.ID, "password")
        password.send_keys("secret_sauce")

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        # find of filter option
        self.visibility_element_located((By.XPATH, "//*[@id='header_container']/div[2]/div/span/select"))
        self.driver.find_element(By.XPATH, "//*[@id='header_container']/div[2]/div/span/select").click()

        # click on the filter for high price of option 
        self.visibility_element_located((By.XPATH, "//*[@id='header_container']/div[2]/div/span/select/option[4]"))
        self.driver.find_element(By.XPATH, "//*[@id='header_container']/div[2]/div/span/select/option[4]").click()

        self.visibility_element_located((By.XPATH, "//*[@id='inventory_container']/div/div[1]/div[2]/div[2]/div"))
        top_price = self.driver.find_element(By.XPATH, "//*[@id='inventory_container']/div/div[1]/div[2]/div[2]/div")
        
        sleep(1)
        self.driver.save_screenshot(self.forlder_path_of_name+f"/high-price-product.png")
        # print(f"price : {top_price}")
        assert top_price

    # Product sayfasındaki fiyatı en yüksek ürünün bulunması testi
    def test_low_product_of_page(self):
        self.visibility_element_located((By.ID, "user-name"))
        user_name = self.driver.find_element(By.ID, "user-name")
        user_name.send_keys("standard_user")

        password = self.driver.find_element(By.ID, "password")
        password.send_keys("secret_sauce")

        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()

        # find of filter option
        self.visibility_element_located((By.XPATH, "//*[@id='header_container']/div[2]/div/span/select"))
        self.driver.find_element(By.XPATH, "//*[@id='header_container']/div[2]/div/span/select").click()

        # click on the filter for low price of option 
        self.visibility_element_located((By.XPATH, "//*[@id='header_container']/div[2]/div/span/select/option[3]"))
        self.driver.find_element(By.XPATH, "//*[@id='header_container']/div[2]/div/span/select/option[3]").click()

        self.visibility_element_located((By.XPATH, "//*[@id='inventory_container']/div/div[1]/div[2]/div[2]/div"))
        low_price = self.driver.find_element(By.XPATH, "//*[@id='inventory_container']/div/div[1]/div[2]/div[2]/div")
        
        sleep(1)
        self.driver.save_screenshot(self.forlder_path_of_name+f"/low-price-product.png")
        print(f"price : {low_price}")
        assert low_price
